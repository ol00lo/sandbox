import openpyxl
from openpyxl.styles import Font
from typing import Dict, List, Tuple
from pathlib import Path
from PyQt6 import QtCore

class BBoxList:
    def __init__(self, output_file: str = "bbox_output.xlsx") -> None:
        self.output_file: str = output_file
        self.bbox_data: Dict[str, List[Tuple[str, QtCore.QRect]]] = {}
    
    def add_bbox(self, bbox: QtCore.QRect, label: str, image_path: str):
        if image_path not in self.bbox_data:
            self.bbox_data[image_path] = []
        self.bbox_data[image_path].append((label, bbox))
    
    def save(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bounding Boxes"

        headers = ["Path", "n_boxes", "Label", "x1", "y1", "x2", "y2"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        row_num = 2

        for image_path, bboxes in self.bbox_data.items():
            n_boxes = len(bboxes)

            for label, bbox in bboxes:
                sheet.cell(row=row_num, column=1, value=str(image_path))
                sheet.cell(row=row_num, column=2, value=n_boxes)
                sheet.cell(row=row_num, column=3, value=str(label))
                sheet.cell(row=row_num, column=4, value=int(bbox.left()))
                sheet.cell(row=row_num, column=5, value=int(bbox.top()))
                sheet.cell(row=row_num, column=6, value=int(bbox.right()))
                sheet.cell(row=row_num, column=7, value=int(bbox.bottom()))
                
                row_num += 1

        workbook.save(self.output_file)
        print(f"Xlsx saved to: {Path(self.output_file).absolute()}")


if __name__ :
    bbox_list = BBoxList()

    rect1 = QtCore.QRect(10, 20, 30, 40)
    rect2 = QtCore.QRect(50, 60, 70, 80)
    rect3 = QtCore.QRect(15, 25, 35, 45)

    bbox_list.add_bbox(rect1, "cat", "images/img1.jpg")
    bbox_list.add_bbox(rect2, "dog", "images/img1.jpg")
    bbox_list.add_bbox(rect3, "cat", "images/img2.jpg")

    bbox_list.save()